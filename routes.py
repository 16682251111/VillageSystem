"""
村户信息管理系统 - API 路由
提供户/成员的 CRUD、Excel 导入、搜索筛选等接口。
"""

import os
import re
import uuid
from datetime import datetime

from flask import Blueprint, request, jsonify, current_app
from werkzeug.utils import secure_filename

from models import db, Household, Member
from config import ALLOWED_EXTENSIONS, HOUSEHOLD_EXCEL_COLUMNS, MEMBER_EXCEL_COLUMNS

api = Blueprint("api", __name__, url_prefix="/api")


# ==================== 工具函数 ====================

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_photo(file):
    """保存照片，返回访问路径"""
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit(".", 1)[1].lower()
        new_name = f"{uuid.uuid4().hex}.{ext}"
        photo_dir = current_app.config.get("PHOTO_FOLDER")
        os.makedirs(photo_dir, exist_ok=True)
        filepath = os.path.join(photo_dir, new_name)
        file.save(filepath)
        return f"/uploads/photos/{new_name}"
    return ""


# ==================== 户 CRUD ====================

@api.route("/households", methods=["GET"])
def list_households():
    """获取户列表，支持搜索、筛选、分页"""
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    keyword = request.args.get("keyword", "").strip()
    house_type = request.args.get("house_type", "").strip()
    group_name = request.args.get("group_name", "").strip()

    query = Household.query

    if keyword:
        pattern = f"%{keyword}%"
        query = query.outerjoin(Member).filter(
            db.or_(
                Household.householder_name.like(pattern),
                Household.house_number.like(pattern),
                Household.householder_phone.like(pattern),
                Member.name.like(pattern),
                Member.phone.like(pattern),
            )
        ).distinct()

    if house_type:
        query = query.filter(Household.house_type == house_type)
    if group_name:
        query = query.filter(Household.group_name == group_name)

    query = query.order_by(Household.house_number)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return jsonify({
        "code": 0,
        "data": {
            "items": [h.to_dict() for h in pagination.items],
            "total": pagination.total,
            "page": pagination.page,
            "pages": pagination.pages,
            "per_page": pagination.per_page,
        }
    })


@api.route("/households/<int:hh_id>", methods=["GET"])
def get_household(hh_id):
    """获取单个户详情（含成员）"""
    hh = Household.query.get_or_404(hh_id)
    return jsonify({"code": 0, "data": hh.to_dict(include_members=True)})


@api.route("/households", methods=["POST"])
def create_household():
    """新增户"""
    data = request.get_json(silent=True) or {}
    house_number = data.get("house_number", "").strip()
    if not house_number:
        return jsonify({"code": 1, "msg": "户编号不能为空"}), 400

    if Household.query.filter_by(house_number=house_number).first():
        return jsonify({"code": 1, "msg": f"户编号 {house_number} 已存在"}), 400

    hh = Household(
        house_number=house_number,
        householder_name=data.get("householder_name", ""),
        householder_phone=data.get("householder_phone", ""),
        group_name=data.get("group_name", ""),
        house_type=data.get("house_type", "一般户"),
        address=data.get("address", ""),
        latitude=data.get("latitude"),
        longitude=data.get("longitude"),
        photo_path=data.get("photo_path", ""),
        personal_photo=data.get("personal_photo", ""),
        planting=data.get("planting", ""),
        breeding=data.get("breeding", ""),
        notes=data.get("notes", ""),
    )
    db.session.add(hh)
    db.session.commit()
    return jsonify({"code": 0, "data": hh.to_dict(), "msg": "新增成功"})


@api.route("/households/<int:hh_id>", methods=["PUT"])
def update_household(hh_id):
    """编辑户信息"""
    hh = Household.query.get_or_404(hh_id)
    data = request.get_json(silent=True) or {}

    # 如果修改了户编号，检查唯一性
    new_number = data.get("house_number", "").strip()
    if new_number and new_number != hh.house_number:
        if Household.query.filter_by(house_number=new_number).first():
            return jsonify({"code": 1, "msg": f"户编号 {new_number} 已存在"}), 400
        hh.house_number = new_number

    for field in ["householder_name", "householder_phone", "group_name",
                  "house_type", "address", "notes", "photo_path",
                  "personal_photo", "planting", "breeding"]:
        if field in data:
            setattr(hh, field, data[field])
    for field in ["latitude", "longitude"]:
        if field in data:
            val = data[field]
            setattr(hh, field, float(val) if val not in (None, "") else None)

    db.session.commit()
    return jsonify({"code": 0, "data": hh.to_dict(), "msg": "更新成功"})


@api.route("/households/<int:hh_id>", methods=["DELETE"])
def delete_household(hh_id):
    """删除户（级联删除成员）"""
    hh = Household.query.get_or_404(hh_id)
    db.session.delete(hh)
    db.session.commit()
    return jsonify({"code": 0, "msg": "删除成功"})


@api.route("/households/upload-photo", methods=["POST"])
def upload_photo():
    """上传照片"""
    file = request.files.get("file")
    if not file:
        return jsonify({"code": 1, "msg": "未选择文件"}), 400
    path = save_photo(file)
    if not path:
        return jsonify({"code": 1, "msg": "不支持的文件格式"}), 400
    return jsonify({"code": 0, "data": {"url": path}})


# ==================== 成员 CRUD ====================

@api.route("/households/<int:hh_id>/members", methods=["GET"])
def list_members(hh_id):
    """获取某户所有成员"""
    Household.query.get_or_404(hh_id)
    members = Member.query.filter_by(household_id=hh_id).all()
    return jsonify({"code": 0, "data": [m.to_dict() for m in members]})


@api.route("/households/<int:hh_id>/members", methods=["POST"])
def add_member(hh_id):
    """添加成员"""
    Household.query.get_or_404(hh_id)
    data = request.get_json(silent=True) or {}
    if not data.get("name", "").strip():
        return jsonify({"code": 1, "msg": "姓名不能为空"}), 400

    member = Member(
        household_id=hh_id,
        member_code=data.get("member_code", ""),
        name=data.get("name", ""),
        id_card=data.get("id_card", ""),
        phone=data.get("phone", ""),
        relation=data.get("relation", ""),
        gender=data.get("gender", ""),
        birth_date=data.get("birth_date", ""),
        notes=data.get("notes", ""),
    )
    db.session.add(member)
    db.session.commit()
    return jsonify({"code": 0, "data": member.to_dict(), "msg": "添加成功"})


@api.route("/members/<int:m_id>", methods=["PUT"])
def update_member(m_id):
    """编辑成员"""
    member = Member.query.get_or_404(m_id)
    data = request.get_json(silent=True) or {}
    for field in ["member_code", "name", "id_card", "phone", "relation", "gender", "birth_date", "notes"]:
        if field in data:
            setattr(member, field, data[field])
    db.session.commit()
    return jsonify({"code": 0, "data": member.to_dict(), "msg": "更新成功"})


@api.route("/members/<int:m_id>", methods=["DELETE"])
def delete_member(m_id):
    """删除成员"""
    member = Member.query.get_or_404(m_id)
    db.session.delete(member)
    db.session.commit()
    return jsonify({"code": 0, "msg": "删除成功"})


# ==================== 统计接口 ====================

@api.route("/stats", methods=["GET"])
def get_stats():
    """获取统计数据"""
    total = Household.query.count()
    total_members = Member.query.count()
    groups = [r[0] for r in db.session.query(Household.group_name).distinct() if r[0]]
    type_counts = {}
    for ht in db.session.query(Household.house_type, db.func.count(Household.id)).group_by(Household.house_type).all():
        if ht[0]:
            type_counts[ht[0]] = ht[1]

    # 有坐标的户（用于地图）
    with_coords = Household.query.filter(
        Household.latitude.isnot(None),
        Household.longitude.isnot(None)
    ).count()

    return jsonify({
        "code": 0,
        "data": {
            "total_households": total,
            "total_members": total_members,
            "groups": groups,
            "type_counts": type_counts,
            "with_coords": with_coords,
        }
    })


@api.route("/map-data", methods=["GET"])
def get_map_data():
    """获取地图标注数据（所有有坐标的户）"""
    households = Household.query.filter(
        Household.latitude.isnot(None),
        Household.longitude.isnot(None)
    ).all()
    return jsonify({
        "code": 0,
        "data": [{
            "id": h.id,
            "house_number": h.house_number,
            "householder_name": h.householder_name,
            "house_type": h.house_type,
            "group_name": h.group_name,
            "latitude": h.latitude,
            "longitude": h.longitude,
            "member_count": h.members.count(),
        } for h in households]
    })


# ==================== Excel 导入 ====================

def _safe_str(val):
    """安全转换为字符串"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _safe_float(val):
    """安全转换为浮点数"""
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _find_column_index_fuzzy(headers, aliases):
    """根据别名列表模糊匹配 Excel 列索引（去除空格、兼容换行）"""
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_clean = str(h).strip().replace("\n", "").replace("\r", "").replace(" ", "")
        for alias in aliases:
            a_clean = alias.strip().replace("\n", "").replace("\r", "").replace(" ", "")
            if h_clean == a_clean or h_clean in a_clean or a_clean in h_clean:
                return i
    return None


def _parse_group_name(sheet_name, first_row_cells):
    """从 sheet 名或第一行解析组名"""
    # 尝试从第一行解析（如 "村：一组" 或 "村：二 组"）
    if first_row_cells:
        first_text = _safe_str(first_row_cells[0])
        import re
        # 先尝试紧密匹配（如"一组"）
        m = re.search(r'[一二三四五六七八九十]+组', first_text)
        if m:
            return m.group()
        # 再尝试带空格的匹配（如"二 组"→"二组"）
        m = re.search(r'[一二三四五六七八九十]+\s+组', first_text)
        if m:
            return m.group().replace(' ', '')
    # 回退：使用 sheet 名
    return str(sheet_name).strip()


def _is_header_row(headers_cells, col_map_target):
    """判断某行是否为表头行（包含序号、户主等关键词）"""
    texts = [_safe_str(c) for c in headers_cells if c is not None]
    combined = ''.join(texts)
    return '序号' in combined or '户主' in combined


@api.route("/import-excel", methods=["POST"])
def import_excel():
    """
    导入 Excel 台账（支持两种格式）：
    格式一（推荐）：每个 sheet 对应一个村民小组，第一行含组名，第二行为表头，数据按序号分行
    格式二（兼容）：单独的户信息表 + 成员信息表
    """
    try:
        import openpyxl
    except ImportError:
        return jsonify({"code": 1, "msg": "请安装 openpyxl: pip install openpyxl"}), 500

    file = request.files.get("file")
    if not file:
        return jsonify({"code": 1, "msg": "未选择文件"}), 400

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls"):
        return jsonify({"code": 1, "msg": "仅支持 .xlsx / .xls 格式"}), 400

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return jsonify({"code": 1, "msg": f"Excel 读取失败: {str(e)}"}), 400

    sheets = list(wb.sheetnames)
    result = {
        "households_created": 0, "households_updated": 0,
        "members_created": 0, "errors": [],
        "sheets_found": sheets,
    }

    # ========== 判断格式类型 ==========
    # 如果所有 sheet 名都像组名（如一组、二组），则使用格式一（每 sheet 一组）
    import re as regex
    is_group_format = all(
        regex.search(r'[一二三四五六七八九十]组', str(sn)) or regex.search(r'^\d+组$', str(sn))
        for sn in sheets
    )

    if is_group_format:
        # ====== 格式一：每 sheet 一组 ======
        for sn in sheets:
            try:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 3:
                    continue

                group_name = _parse_group_name(sn, rows[0])

                header_row_idx = None
                for i in range(min(3, len(rows))):
                    if _is_header_row(rows[i], ['序号', '户主']):
                        header_row_idx = i
                        break

                if header_row_idx is None:
                    result["errors"].append(f"工作表 [{sn}] 未找到表头行，跳过")
                    continue

                headers = rows[header_row_idx]
                data_start = header_row_idx + 1

                hh_col_map = {}
                for field, aliases in HOUSEHOLD_EXCEL_COLUMNS.items():
                    idx = _find_column_index_fuzzy(headers, aliases)
                    if idx is not None:
                        hh_col_map[field] = idx

                # 如果 seq 和 house_number 指向同一列，尝试为 house_number 寻找「户主」列
                if ("seq" in hh_col_map and "house_number" in hh_col_map
                        and hh_col_map["seq"] == hh_col_map["house_number"]):
                    alt_idx = _find_column_index_fuzzy(headers, ["户主编码", "户主", "户编号", "家庭编号", "农户编号", "户号"])
                    if alt_idx is not None and alt_idx != hh_col_map["seq"]:
                        hh_col_map["house_number"] = alt_idx

                m_col_map = {}
                for field, aliases in MEMBER_EXCEL_COLUMNS.items():
                    idx = _find_column_index_fuzzy(headers, aliases)
                    if idx is not None:
                        m_col_map[field] = idx

                if "house_number" not in hh_col_map:
                    result["errors"].append(f"工作表 [{sn}] 缺少「户主」列，跳过")
                    continue

                current_hh = None
                sheet_hh_created = 0
                sheet_hh_updated = 0
                sheet_members = 0

                for row_idx in range(data_start, len(rows)):
                    row = rows[row_idx]
                    if all(v is None for v in row):
                        continue

                    seq_idx = hh_col_map.get("seq", hh_col_map.get("house_number"))
                    seq_val = _safe_str(row[seq_idx]) if seq_idx is not None else ""
                    household_code = _safe_str(row[hh_col_map["house_number"]]) if "house_number" in hh_col_map else ""
                    member_code = _safe_str(row[m_col_map["member_code"]]) if "member_code" in m_col_map else ""

                    is_house_row = bool(household_code) and not bool(member_code) and bool(seq_val or household_code)

                    if is_house_row:
                        if current_hh is not None:
                            db.session.flush()

                        address = _safe_str(row[hh_col_map["address"]]) if "address" in hh_col_map else ""
                        householder_name = ""
                        if "householder_name" in hh_col_map:
                            householder_name = _safe_str(row[hh_col_map["householder_name"]])
                        if not householder_name:
                            householder_name = household_code

                        # 使用 no_autoflush 避免查询时自动 flush 未就绪的对象
                        with db.session.no_autoflush:
                            hh = None
                            if group_name and address:
                                hh = Household.query.filter_by(group_name=group_name, address=address).first()
                            if not hh and household_code and group_name:
                                hh = Household.query.filter_by(house_number=household_code, group_name=group_name).first()

                        is_new_hh = hh is None
                        if is_new_hh:
                            hh = Household(house_number=household_code)
                            sheet_hh_created += 1
                        else:
                            sheet_hh_updated += 1

                        # 先设置所有字段
                        hh.householder_name = householder_name or hh.householder_name
                        hh.group_name = group_name
                        for fname, condition in [
                            ("householder_phone", "householder_phone" in hh_col_map),
                            ("address", "address" in hh_col_map),
                            ("house_type", "house_type" in hh_col_map),
                        ]:
                            if condition:
                                val = _safe_str(row[hh_col_map[fname]])
                                if val and val != "None":
                                    setattr(hh, fname, val)

                        for fname, col_key in [
                            ("photo_path", "house_photo"),
                            ("personal_photo", "personal_photo"),
                            ("planting", "planting"),
                            ("breeding", "breeding"),
                            ("notes", "notes"),
                        ]:
                            if col_key in hh_col_map:
                                val = _safe_str(row[hh_col_map[col_key]])
                                if val and val != "None":
                                    setattr(hh, fname, val)

                        if is_new_hh:
                            db.session.add(hh)
                            db.session.flush()  # flush 以获取 id，供成员行使用
                        current_hh = hh

                    elif current_hh is not None and member_code:
                        member_name = _safe_str(row[m_col_map["name"]]) if "name" in m_col_map else member_code
                        member = Member(
                            household_id=current_hh.id,
                            member_code=member_code,
                            name=member_name or member_code,
                        )
                        for fname, col_key in [
                            ("relation", "relation"),
                            ("id_card", "id_card"),
                            ("gender", "gender"),
                            ("phone", "phone"),
                            ("notes", "notes"),
                        ]:
                            if col_key in m_col_map:
                                val = _safe_str(row[m_col_map[col_key]])
                                if val and val != "None":
                                    setattr(member, fname, val)

                        db.session.add(member)
                        sheet_members += 1

                db.session.flush()
                result["households_created"] += sheet_hh_created
                result["households_updated"] += sheet_hh_updated
                result["members_created"] += sheet_members

            except Exception as e:
                db.session.rollback()
                result["errors"].append(f"工作表 [{sn}] 处理异常: {str(e)}")

    else:
        # ====== 格式二（兼容旧版）：户信息 + 成员信息 两个工作表 ======
        result["errors"].append("Excel 为旧格式（户信息+成员信息双表模式），请使用多 sheet 格式。")
        # 如果你需要保留旧格式兼容，在此处添加旧版逻辑

    wb.close()
    db.session.commit()
    return jsonify({
        "code": 0,
        "data": result,
        "msg": f"导入完成：新增 {result['households_created']} 户，更新 {result['households_updated']} 户，"
               f"新增成员 {result['members_created']} 人"
               + (f"，{len(result['errors'])} 条警告" if result['errors'] else "")
    })
